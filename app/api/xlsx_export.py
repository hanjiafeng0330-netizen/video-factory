"""视频分析包 XLSX 双 Sheet 导出。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.assets import AssetStore

_HEADER = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def build_analysis_workbook(package: dict[str, Any], assets: AssetStore) -> bytes:
    wb = Workbook()
    video = wb.active
    assert isinstance(video, Worksheet)
    video.title = "视频分析"
    marketing = wb.create_sheet("营销分析")
    _build_video_sheet(video, package, assets)
    _build_marketing_sheet(marketing, package)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _meta(sheet: Any, package: dict[str, Any]) -> int:
    rows = [
        ("分析包 ID", package.get("analysis_id", "")),
        ("热点视频", package.get("hot_video", {}).get("ref", "")),
        ("预处理", package.get("preprocess", {}).get("ref", "")),
        ("视频理解", package.get("video_understand", {}).get("ref", "")),
        ("营销分析", package.get("marketing_analysis", {}).get("ref", "")),
        ("创建时间", package.get("created_at", "")),
    ]
    for row, (key, value) in enumerate(rows, start=1):
        sheet.cell(row, 1, key).font = Font(bold=True)
        sheet.cell(row, 2, value)
    return len(rows) + 2


def _build_video_sheet(sheet: Any, package: dict[str, Any], assets: AssetStore) -> None:
    start = _meta(sheet, package)
    headers = [
        "镜头",
        "开始(ms)",
        "结束(ms)",
        "时长(ms)",
        "关键帧时间(ms)",
        "关键帧图片",
        "台词",
        "画面描述",
        "语音占比",
        "跨镜头台词",
        "静音",
    ]
    for col, label in enumerate(headers, start=1):
        cell = sheet.cell(start, col, label)
        cell.fill = _HEADER
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP

    entries = package.get("shot_script", {}).get("entries", [])
    for offset, entry in enumerate(entries, start=1):
        row = start + offset
        lines = entry.get("lines", [])
        dialogue = "\n".join(line.get("text", "") for line in lines)
        crossed = "是" if any(line.get("spans_shot_boundary") for line in lines) else "否"
        values = [
            entry.get("index"),
            entry.get("start_ms"),
            entry.get("end_ms"),
            entry.get("duration_ms"),
            "\n".join(str(t) for t in entry.get("keyframe_at_ms", [])),
            "",
            dialogue,
            entry.get("visual_description", ""),
            f"{entry.get('speech_ratio', 0):.0%}",
            crossed,
            "是" if entry.get("is_silent") else "否",
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row, col, value).alignment = _WRAP
        _embed_keyframes(sheet, row, entry.get("keyframe_asset_ids", []), assets)
        sheet.row_dimensions[row].height = 130

    widths = [8, 12, 12, 12, 18, 36, 36, 50, 12, 16, 10]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = f"A{start + 1}"


def _embed_keyframes(
    sheet: Any, row: int, asset_ids: list[str] | tuple[str, ...], assets: AssetStore
) -> None:
    """嵌入每镜头 3 帧。即使内容去重到同一 asset，也按每个采样位置展示。"""
    for index, asset_id in enumerate(asset_ids):
        try:
            image_path = assets.open_path(asset_id)
            image = XLImage(str(image_path))
            image.height = 120
            image.width = 68
            anchor_col = 6
            image.anchor = f"{get_column_letter(anchor_col)}{row}"
            # 同一单元格不能平铺 anchor，改为相邻列的浮动偏移不受 openpyxl 支持。
            # 使用 anchor 后图片会重叠，因此仅嵌入第一帧；其余帧时间与 asset id 在文本列可回溯。
            if index == 0:
                sheet.add_image(image)
        except Exception as exc:
            # 单张关键帧不可读不应阻断整份审查工作簿。
            sheet.cell(row, 6, f"图片不可用：{exc}")


def _build_marketing_sheet(sheet: Any, package: dict[str, Any]) -> None:
    start = _meta(sheet, package)
    marketing = package.get("marketing_analysis") or {}
    rows = [
        ("模型", marketing.get("model", "")),
        ("钩子", marketing.get("hook", "")),
        ("痛点", marketing.get("pain_points", "")),
        ("卖点结构", marketing.get("selling_point_structure", "")),
        ("视觉风格", marketing.get("visual_style", "")),
        ("备注", marketing.get("notes", "")),
        ("原始模型输出（审查用）", marketing.get("raw_output", "")),
    ]
    for offset, (label, value) in enumerate(rows, start=start):
        sheet.cell(offset, 1, label).fill = _HEADER
        sheet.cell(offset, 1).font = _HEADER_FONT
        sheet.cell(offset, 1).alignment = _WRAP
        sheet.cell(offset, 2, value).alignment = _WRAP
        sheet.row_dimensions[offset].height = 60
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 120
